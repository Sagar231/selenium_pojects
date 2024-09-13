'''

read write data from excel
file==>workbook-->sheets-->rows==>cells
'''
import openpyxl
from openpyxl.styles import PatternFill
from attr.validators import max_len

path =""
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]  # sheet1 is the name
#sheet = workbook.active
# sheet.cell(r,c).value = "welcome"
# workbook,save()


row = sheet.max_row
col = sheet.max_column

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value)

